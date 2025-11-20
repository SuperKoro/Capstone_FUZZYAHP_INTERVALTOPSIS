"""
Excel Handler Module
Handles Excel template generation and import/export operations
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Tuple
import os


class ExcelHandler:
    """Handles all Excel-related operations"""
    
    @staticmethod
    def generate_ahp_template(criteria_names: List[str], expert_name: str, 
                             output_path: str) -> None:
        """
        Generate Excel template for AHP pairwise comparisons
        
        Args:
            criteria_names: List of criterion names
            expert_name: Name of the expert
            output_path: Path to save the Excel file
        """
        n = len(criteria_names)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AHP Comparisons"
        
        # Styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = f"Fuzzy AHP Pairwise Comparison Template - Expert: {expert_name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Instructions
        ws.merge_cells('A2:D2')
        ws['A2'] = "Instructions: Enter scale values from -9 to 9 (excluding 0) for each comparison"
        ws['A2'].font = Font(italic=True)
        
        # Scale reference
        ws['A4'] = "Scale Reference:"
        ws['A4'].font = Font(bold=True)
        
        scale_ref = [
            "9 = Absolutely more important",
            "7 = Strongly very more important",
            "5 = Strongly more important",
            "3 = Moderately more important",
            "1 = Equally important",
            "-3 = Moderately less important",
            "-5 = Strongly less important",
            "-7 = Strongly very less important",
            "-9 = Absolutely less important"
        ]
        
        for idx, ref in enumerate(scale_ref, start=5):
            ws[f'A{idx}'] = ref
        
        # Comparison matrix header
        start_row = 16
        ws[f'A{start_row}'] = "Criterion 1"
        ws[f'B{start_row}'] = "Criterion 2"
        ws[f'C{start_row}'] = "Scale Value"
        ws[f'D{start_row}'] = "Interpretation"
        
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{start_row}']
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Generate all pairwise comparisons
        row = start_row + 1
        for i in range(n):
            for j in range(i + 1, n):
                ws[f'A{row}'] = criteria_names[i]
                ws[f'B{row}'] = criteria_names[j]
                ws[f'C{row}'] = ""  # To be filled by expert
                ws[f'D{row}'] = ""  # Will show interpretation
                
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].border = border
                
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 35
        
        # Save workbook
        wb.save(output_path)
    
    @staticmethod
    def import_ahp_comparisons(file_path: str, criteria_mapping: Dict[str, int]) -> List[Dict]:
        """
        Import AHP comparisons from Excel file
        
        Args:
            file_path: Path to the Excel file
            criteria_mapping: Dictionary mapping criterion names to IDs
            
        Returns:
            List of comparison dictionaries
        """
        df = pd.read_excel(file_path, sheet_name="AHP Comparisons", skiprows=15)
        
        comparisons = []
        
        for _, row in df.iterrows():
            criterion1_name = str(row['Criterion 1']).strip()
            criterion2_name = str(row['Criterion 2']).strip()
            scale_value = row['Scale Value']
            
            # Skip empty rows
            if pd.isna(scale_value) or criterion1_name == 'nan' or criterion2_name == 'nan':
                continue
            
            # Validate scale value
            scale_value = int(scale_value)
            if scale_value == 0 or scale_value < -9 or scale_value > 9:
                raise ValueError(f"Invalid scale value: {scale_value}")
            
            # Get criterion IDs
            if criterion1_name not in criteria_mapping or criterion2_name not in criteria_mapping:
                raise ValueError(f"Unknown criterion: {criterion1_name} or {criterion2_name}")
            
            criterion1_id = criteria_mapping[criterion1_name]
            criterion2_id = criteria_mapping[criterion2_name]
            
            # Get fuzzy values
            from algorithms.fuzzy_ahp import FuzzyAHP
            fuzzy_l, fuzzy_m, fuzzy_u = FuzzyAHP.get_fuzzy_number(scale_value)
            
            comparisons.append({
                'criterion1_id': criterion1_id,
                'criterion2_id': criterion2_id,
                'fuzzy_l': fuzzy_l,
                'fuzzy_m': fuzzy_m,
                'fuzzy_u': fuzzy_u
            })
        
        return comparisons
    
    @staticmethod
    def export_results(project_info: Dict, criteria: List[Dict], alternatives: List[Dict],
                      weights: Dict, topsis_results: Dict, output_path: str) -> None:
        """
        Export comprehensive results to Excel
        
        Args:
            project_info: Project metadata
            criteria: List of criteria with weights
            alternatives: List of alternatives
            weights: AHP weights and CR
            topsis_results: TOPSIS calculation results
            output_path: Path to save the Excel file
        """
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # Sheet 1: Project Summary
            summary_data = {
                'Project Name': [project_info['name']],
                'Description': [project_info.get('description', '')],
                'Created Date': [project_info.get('created_date', '')],
                'Modified Date': [project_info.get('modified_date', '')]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Project Summary', index=False)
            
            # Sheet 2: Criteria Weights
            criteria_data = {
                'Criterion': [c['name'] for c in criteria],
                'Weight': [c['weight'] for c in criteria],
                'Type': ['Benefit' if c['is_benefit'] else 'Cost' for c in criteria]
            }
            df_criteria = pd.DataFrame(criteria_data)
            df_criteria.to_excel(writer, sheet_name='Criteria Weights', index=False)
            
            # Add CR information
            ws_criteria = writer.sheets['Criteria Weights']
            ws_criteria['A' + str(len(criteria) + 4)] = "Consistency Ratio (CR):"
            ws_criteria['B' + str(len(criteria) + 4)] = weights.get('cr', 0)
            ws_criteria['A' + str(len(criteria) + 5)] = "CR Status:"
            cr_status = "Acceptable (< 0.1)" if weights.get('cr', 0) < 0.1 else "Needs Review (>= 0.1)"
            ws_criteria['B' + str(len(criteria) + 5)] = cr_status
            
            # Sheet 3: Final Rankings
            ranking_data = {
                'Rank': list(range(1, len(alternatives) + 1)),
                'Alternative': [alternatives[i]['name'] for i in topsis_results['ranking']],
                'Closeness Coefficient': [topsis_results['closeness_coefficients'][i] 
                                         for i in topsis_results['ranking']],
                'Distance to PIS': [topsis_results['distances_to_PIS'][i] 
                                   for i in topsis_results['ranking']],
                'Distance to NIS': [topsis_results['distances_to_NIS'][i] 
                                   for i in topsis_results['ranking']]
            }
            df_ranking = pd.DataFrame(ranking_data)
            df_ranking.to_excel(writer, sheet_name='Final Rankings', index=False)
            
            # Format the workbook
            workbook = writer.book
            
            # Format headers
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
